import os
import pandas as pd
from typing import List, Optional, cast # Importiere cast

# openpyxl Imports erweitert, um Proxy-Typen zu vermeiden und MergedCell zu behandeln
from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell # MergedCell importieren
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import Color # Color-Typ importieren, falls benötigt

def _is_cell_yellow(cell: Cell) -> bool:
    """Return True if the openpyxl cell appears to have a yellow fill.
    Returns:
        bool: True if the cell is filled with yellow color, False otherwise.
    """
    try:
        # Testing the fill of the cell
        fill: PatternFill = cast(PatternFill, cell.fill)

        if not fill or getattr(fill, 'patternType', None) is None:
            return False

        fg = getattr(fill, 'fgColor', None)
        if fg is None:
            return False
            
        rgb: Optional[str] = None
        
        # ... Logic for RGB colors ...
        if getattr(fg, 'type', None) == 'rgb' and getattr(fg, 'rgb', None):
            rgb = str(fg.rgb)
        elif getattr(fg, 'rgb', None):
            rgb = str(fg.rgb)
        elif getattr(fg, 'indexed', None) is not None:
            try:
                # 1. Ensure fg.indexed is not None
                idx_val = fg.indexed
                if idx_val is not None:
                    # 2. Cast to int after converting to str to avoid Hashable issues
                    idx = int(str(idx_val)) # <-- Addition of str() solves the hashable issue
                    if idx in (5, 6, 7):
                        return True
            except Exception:
                pass

        if not rgb:
            return False
            
        rgb_norm = rgb.upper()
        if rgb_norm.startswith('00') and len(rgb_norm) == 8:
            rgb_norm = rgb_norm[2:]
            
        return rgb_norm.endswith('FFFF00') or rgb_norm == 'FFFF00'
    except Exception:
        return False


def read_excel_file() -> pd.DataFrame:
    """Read the Excel file from `Excel/data.xlsx` and detect PO cells highlighted in yellow.
    Returns:
        pd.DataFrame: DataFrame containing the Excel data with an additional 'PO_HIGHLIGHTED' column.
    """
    #-- CONFIGURATION --
    EXCEL_FILE_PATH = os.path.join("Excel", "data.xlsx")

    print("\nℹ️\tStep 1: Reading Excel Data File...\n")
    try:
        df_excel = pd.read_excel(EXCEL_FILE_PATH)
        df_excel = df_excel[[ 'as mentioned in kolom M, N', 'PO', 'PCME2', 'PCME3']]
        df_excel = df_excel.rename(columns={'as mentioned in kolom M, N': 'MCODE'})
        df_excel['PO_HIGHLIGHTED'] = False

        print("ℹ️\tStep 2: Loading Workbook styles to detect yellow highlights...")
        
        # Load workbook and active worksheet with styles
        wb: Workbook = load_workbook(EXCEL_FILE_PATH, data_only=False)
        ws: Worksheet = cast(Worksheet, wb.active)

        header_row: Optional[int] = None
        po_col_idx: Optional[int] = None

        # Find Header row and PO column index
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell_item in row:
                
                cell: Cell = cast(Cell, cell_item)
                
                if cell.value == 'PO':
                    header_row = cell.row
                    po_col_idx = cell.column
                    break
            if header_row is not None:
                break

        # Fallback Logic 
        if po_col_idx is None:
            headers: List[str] = list(df_excel.columns)
            try:
                pandas_po_idx: int = headers.index('PO')
                po_col_idx = pandas_po_idx + 1
                header_row = 1
            except ValueError:
                po_col_idx = None

        if po_col_idx is None:
            print("⚠️\tCould not determine PO column in the spreadsheet; skipping highlight detection.")
            return df_excel

        final_header_row: int = header_row if header_row is not None else 1
        
        for df_idx, _ in df_excel.iterrows():
            # **KORREKTUR DER FEHLERHAFTEN ZEILE**
            # Wir casten df_idx explizit zu int, um Pylance zu beruhigen,
            # da wir wissen, dass der Index 0-basiert ist und für die Zeilenberechnung verwendet wird.
            df_index_pos: int = cast(int, df_idx)
            
            # Die Zeilennummer in Excel ist header + 1 + 0-basierter Index
            excel_row: int = final_header_row + 1 + df_index_pos 
            
            try:
                if po_col_idx is not None:
                    cell_to_check_item = ws.cell(row=excel_row, column=po_col_idx)
                    
                    if isinstance(cell_to_check_item, MergedCell):
                        continue 
                    
                    cell_to_check: Cell = cast(Cell, cell_to_check_item)
                    
                    if _is_cell_yellow(cell_to_check):
                        # Fehler 5 wird mit dem expliziten Cast von df_idx zu int behoben:
                        df_excel.at[df_index_pos, 'PO_HIGHLIGHTED'] = True
            except Exception:
                continue

        print("✅\tSuccessfully read Excel data and detected highlights.")
        return df_excel
        
    except Exception as e:
        raise RuntimeError(f"Failed to read Excel file {EXCEL_FILE_PATH}") from e